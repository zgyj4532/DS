from services.finance_service import FinanceService
from fastapi import APIRouter, HTTPException, Query
from pydantic import BaseModel
from typing import Optional, List, Dict, Any, cast
from core.config import Settings, settings
from core.database import get_conn
from services.finance_service import split_order_funds
from core.config import VALID_PAY_WAYS, POINTS_DISCOUNT_RATE
from core.table_access import build_dynamic_select, get_table_structure, _quote_identifier
from decimal import Decimal
import uuid
from datetime import datetime, timedelta
from enum import Enum
import json
import threading
import time
from core.logging import get_logger
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import List, Dict, Any
from fastapi.responses import StreamingResponse

# ==================== 新增：导入 Redis 用于分布式锁 ====================
import redis
import redis.exceptions

# ==================== 新增：导入微信发货管理模块 ====================
from .wechat_shipping import WechatShippingManager

logger = get_logger(__name__)
router = APIRouter()


# ==================== 新增：Redis 客户端初始化（带容错） ====================
def _get_redis_client():
    """获取 Redis 客户端，如果未配置则返回 None"""
    try:
        # 尝试从 settings 获取 Redis 配置，如果没有则使用默认本地配置
        redis_host = getattr(settings, 'REDIS_HOST', 'localhost')
        redis_port = getattr(settings, 'REDIS_PORT', 6379)
        redis_db = getattr(settings, 'REDIS_DB', 0)

        client = redis.Redis(
            host=redis_host,
            port=redis_port,
            db=redis_db,
            decode_responses=True,
            socket_connect_timeout=2,
            socket_timeout=2
        )
        # 测试连接
        client.ping()
        return client
    except Exception as e:
        logger.warning(f"Redis 连接失败（将使用数据库兜底）: {e}")
        return None


# 全局 Redis 客户端
redis_client = _get_redis_client()


# 在 order.py 的 _cancel_expire_orders 函数中
def _cancel_expire_orders():
    """每分钟扫描一次，把过期的 pending_pay 订单取消"""
    while True:
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    now = datetime.now()
                    cur.execute("""
                        SELECT id, order_number
                        FROM orders
                        WHERE status='pending_pay'
                          AND expire_at IS NOT NULL
                          AND expire_at <= %s
                    """, (now,))
                    for o in cur.fetchall():
                        oid, ono = o["id"], o["order_number"]

                        # ✅ 新增：删除该订单的待发放奖励记录
                        cur.execute(
                            "SELECT id FROM pending_rewards WHERE order_id = %s AND status = 'pending'",
                            (oid,)
                        )
                        rewards = cur.fetchall()
                        for reward in rewards:
                            cur.execute(
                                "DELETE FROM pending_rewards WHERE id = %s",
                                (reward['id'],)
                            )
                            print(f"[expire] 删除订单 {ono} 的待发放奖励记录: ID={reward['id']}")

                        # 回滚库存
                        cur.execute(
                            "SELECT product_id,quantity FROM order_items WHERE order_id=%s",
                            (oid,)
                        )
                        for it in cur.fetchall():
                            cur.execute(
                                "UPDATE product_skus SET stock=stock+%s WHERE product_id=%s",
                                (it["quantity"], it["product_id"])
                            )

                        # 改状态
                        cur.execute(
                            "UPDATE orders SET status='cancelled',updated_at=NOW() WHERE id=%s",
                            (oid,)
                        )
                        print(f"[expire] 订单 {ono} 已自动取消")
                    conn.commit()
        except Exception as e:
            print(f"[expire] error: {e}")
        time.sleep(60)


def start_order_expire_task():
    """由 api.order 包初始化时调用一次即可"""
    t = threading.Thread(target=_cancel_expire_orders, daemon=True)
    t.start()
    print("[expire] 订单过期守护线程已启动")


# ==================== 新增：定时同步微信订单状态（解决资金结算问题） ====================
def _sync_wechat_order_status():
    """
    定时同步微信订单状态（每30分钟执行一次）
    解决用户通过微信确认收货组件确认收货后，后端状态未更新的问题
    """
    while True:
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    # 查询已发货(pending_recv)且已同步到微信(wechat_shipping_status=1)的订单
                    # 且最近1小时内未同步过的订单
                    cur.execute("""
                        SELECT id, order_number, transaction_id, status, user_id
                        FROM orders
                        WHERE status='pending_recv'
                          AND wechat_shipping_status = 1
                          AND transaction_id IS NOT NULL
                          AND (
                              wechat_last_sync_time IS NULL 
                              OR wechat_last_sync_time <= DATE_SUB(NOW(), INTERVAL 1 HOUR)
                          )
                        LIMIT 50
                    """)

                    orders = cur.fetchall()

                    for order in orders:
                        try:
                            # 查询微信侧的订单状态
                            wx_result = WechatShippingManager.get_order(order['transaction_id'])

                            # 更新同步时间
                            cur.execute(
                                "UPDATE orders SET wechat_last_sync_time = NOW() WHERE id = %s",
                                (order['id'],)
                            )

                            # 修复：errcode 转为字符串比较，防止微信返回字符串 "0"
                            if str(wx_result.get('errcode', '')) == '0':
                                # 修复：兼容 order_state 在顶层或嵌套在 order 对象中的情况，并转为整数
                                raw_state = wx_result.get('order_state') or (wx_result.get('order') or {}).get(
                                    'order_state')
                                try:
                                    order_state = int(raw_state) if raw_state is not None else None
                                except (ValueError, TypeError):
                                    order_state = None
                                    logger.warning(f"[wx_sync] order_state 转换失败，原始值: {raw_state}")

                                # 微信状态：1待发货 2已发货 3确认收货 4交易完成 5已退款
                                if order_state == 3:
                                    # 用户已确认收货，更新订单状态为已完成
                                    logger.info(
                                        f"[wx_sync] 订单 {order['order_number']} 微信状态已确认收货，更新本地状态")

                                    # 更新订单状态
                                    cur.execute("""
                                        UPDATE orders 
                                        SET status='completed', 
                                            completed_at=NOW(),
                                            updated_at=NOW()
                                        WHERE id=%s
                                    """, (order['id'],))

                                    # 记录日志
                                    cur.execute("""
                                        INSERT INTO wechat_shipping_logs 
                                        (order_id, order_number, transaction_id, action_type, is_success, response_data, created_at)
                                        VALUES (%s, %s, %s, 'sync', 1, %s, NOW())
                                    """, (
                                        order['id'],
                                        order['order_number'],
                                        order['transaction_id'],
                                        json.dumps(wx_result)
                                    ))

                                    # 可选：触发资金结算（如果之前未结算）
                                    try:
                                        fs = FinanceService()
                                        # 这里可以调用资金结算逻辑，如果之前未在支付时结算的话
                                        # 注意：根据业务逻辑，资金可能已经在支付时拆分，这里只是更新状态
                                    except Exception as e:
                                        logger.error(f"[wx_sync] 订单 {order['order_number']} 资金结算异常: {e}")

                                    conn.commit()
                                    logger.info(f"[wx_sync] 订单 {order['order_number']} 状态已同步为完成")

                                elif order_state == 4:
                                    # 交易完成（可能已过确认收货期）
                                    cur.execute("""
                                        UPDATE orders 
                                        SET status='completed', 
                                            completed_at=NOW(),
                                            updated_at=NOW()
                                        WHERE id=%s
                                    """, (order['id'],))

                                    # 补充：记录日志
                                    cur.execute("""
                                        INSERT INTO wechat_shipping_logs 
                                        (order_id, order_number, transaction_id, action_type, is_success, remark, response_data, created_at)
                                        VALUES (%s, %s, %s, 'sync', 1, %s, %s, NOW())
                                    """, (
                                        order['id'],
                                        order['order_number'],
                                        order['transaction_id'],
                                        "微信状态同步：交易完成(4)",
                                        json.dumps(wx_result)
                                    ))
                                    conn.commit()
                                    logger.info(f"[wx_sync] 订单 {order['order_number']} 状态已同步为完成(交易完成)")

                            else:
                                # 查询失败记录日志
                                logger.warning(
                                    f"[wx_sync] 查询订单 {order['order_number']} 微信状态失败: {wx_result.get('errmsg')}")
                                cur.execute("""
                                    INSERT INTO wechat_shipping_logs 
                                    (order_id, order_number, transaction_id, action_type, is_success, response_data, created_at)
                                    VALUES (%s, %s, %s, 'sync', 0, %s, NOW())
                                """, (
                                    order['id'],
                                    order['order_number'],
                                    order['transaction_id'],
                                    json.dumps(wx_result)
                                ))
                                conn.commit()

                        except Exception as e:
                            logger.error(f"[wx_sync] 同步订单 {order['order_number']} 状态异常: {e}", exc_info=True)
                            conn.rollback()

        except Exception as e:
            logger.error(f"[wx_sync] 定时同步任务异常: {e}", exc_info=True)

        time.sleep(1800)  # 30分钟执行一次


def start_wechat_status_sync_task():
    """启动微信订单状态同步守护线程"""
    t = threading.Thread(target=_sync_wechat_order_status, daemon=True)
    t.start()
    logger.info("[wx_sync] 微信订单状态同步守护线程已启动（每30分钟同步一次）")


class OrderManager:
    @staticmethod
    def _build_orders_select(cursor) -> str:
        structure = get_table_structure(cursor, "orders")
        select_parts = []
        for field in structure['fields']:
            if field in structure['asset_fields']:
                select_parts.append(f"COALESCE({_quote_identifier(field)}, 0) AS {_quote_identifier(field)}")
            else:
                select_parts.append(_quote_identifier(field))
        return ", ".join(select_parts)

    @staticmethod
    def create(
            user_id: int,
            address_id: Optional[int],
            custom_addr: Optional[dict],
            specifications: Optional[str] = None,
            buy_now: bool = False,
            buy_now_items: Optional[List[Dict[str, Any]]] = None,
            delivery_way: str = "platform",
            points_to_use: Optional[Decimal] = None,
            coupon_id: Optional[int] = None,
            # ==================== 新增：幂等性 Key（前端生成，用于防重放） ====================
            idempotency_key: Optional[str] = None,
            # ==================== 新增：商家ID（可选，默认0=平台自营） ====================
            merchant_id: Optional[int] = None
    ) -> Optional[str]:
        """
        创建订单（已增加幂等性校验，防止重复创建，支持多商家订单）

        幂等性策略：
        1. Redis 分布式锁：防止并发重复提交（5秒锁）
        2. 业务层幂等检查：检查最近1分钟内是否有未取消的订单
        3. 数据库唯一索引：最后一道防线（order_number 唯一）

        商家处理逻辑：
        - 如果提供了 merchant_id，直接使用该值
        - 如果未提供，从 buy_now_items 或购物车商品中推断商家ID
        - 如果商品属于不同商家，返回错误（一笔订单只能属于一个商家）
        - 如果无法确定商家，默认使用 0（平台自营）
        """

        # ==================== 新增：Redis 分布式锁 ====================
        lock_key = f"order:create:{user_id}"
        lock_acquired = False

        if redis_client:
            try:
                # NX=True: 只有 key 不存在时才设置成功（获取锁成功）
                # EX=5: 锁 5 秒后自动释放（防止死锁）
                lock_acquired = redis_client.set(lock_key, idempotency_key or "1", nx=True, ex=5)
                if not lock_acquired:
                    logger.warning(f"用户 {user_id} 重复提交订单，Redis 锁拦截")
                    raise HTTPException(
                        status_code=429,
                        detail="订单创建中，请勿重复提交，或等待 5 秒后重试"
                    )
            except redis.exceptions.RedisError as e:
                logger.error(f"Redis 锁操作失败: {e}，将降级为数据库锁")
                # Redis 故障时降级，继续执行（依赖数据库唯一索引兜底）
                lock_acquired = False

        try:
            with get_conn() as conn:
                with conn.cursor() as cur:

                    # ==================== 新增：业务层幂等检查 ====================
                    # 检查该用户最近 1 分钟内是否已有非取消状态的订单（立即购买场景除外）
                    if not buy_now:
                        cur.execute("""
                            SELECT order_number, status, created_at 
                            FROM orders 
                            WHERE user_id = %s 
                              AND created_at > DATE_SUB(NOW(), INTERVAL 1 MINUTE)
                              AND status != 'cancelled'
                            ORDER BY created_at DESC 
                            LIMIT 1
                        """, (user_id,))
                        recent_order = cur.fetchone()

                        if recent_order:
                            logger.warning(
                                f"用户 {user_id} 1 分钟内已有订单 {recent_order['order_number']}，拦截重复创建")
                            raise HTTPException(
                                status_code=400,
                                detail=f"您刚刚已创建订单 {recent_order['order_number']}，请勿重复提交"
                            )

                    # ==================== 新增：幂等 Key 检查（如果提供了的话） ====================
                    if idempotency_key and redis_client:
                        # 检查这个 idempotency_key 是否已经用过（24小时有效期）
                        used_key = f"order:idempotency:{idempotency_key}"
                        if redis_client.exists(used_key):
                            # 返回之前创建的订单号
                            existing_order = redis_client.get(used_key)
                            logger.info(f"幂等 Key 重复，返回已存在订单: {existing_order}")
                            return existing_order

                    # ---------- 1. 组装订单明细 ----------
                    if buy_now:
                        if not buy_now_items:
                            raise HTTPException(status_code=422, detail="立即购买时 buy_now_items 不能为空")
                        items = []
                        product_merchant_ids = set()  # 收集所有商品的商家ID
                        
                        for it in buy_now_items:
                            cur.execute("SELECT is_member_product, user_id FROM products WHERE id = %s", (it["product_id"],))
                            prod = cur.fetchone()
                            if not prod:
                                raise HTTPException(status_code=404,
                                                    detail=f"products 表中不存在 id={it['product_id']}")

                            # 收集商家ID (user_id 即商家ID)
                            product_merchant_ids.add(prod.get("user_id") or 0)

                            sku_id = it.get("sku_id")
                            if not sku_id:
                                cur.execute("SELECT id FROM product_skus WHERE product_id = %s LIMIT 1",
                                            (it['product_id'],))
                                sku_row = cur.fetchone()
                                if sku_row:
                                    sku_id = sku_row.get('id')
                                else:
                                    raise HTTPException(status_code=422,
                                                        detail=f"商品 {it['product_id']} 无可用 SKU，请提供 sku_id")

                            if "price" not in it:
                                raise HTTPException(status_code=422,
                                                    detail=f"buy_now_items 必须包含 price 字段：product_id={it['product_id']}")

                            items.append({
                                "sku_id": sku_id,
                                "product_id": it["product_id"],
                                "quantity": it["quantity"],
                                "price": Decimal(str(it["price"])),
                                "is_vip": prod["is_member_product"]
                            })
                        
                        # 检查商家一致性（一笔订单只能属于一个商家）
                        if len(product_merchant_ids) > 1:
                            raise HTTPException(
                                status_code=400, 
                                detail="一笔订单只能包含同一商家的商品，请分开下单"
                            )
                        
                        # 如果未提供 merchant_id，从商品中推断
                        if merchant_id is None:
                            merchant_id = product_merchant_ids.pop() if product_merchant_ids else 0
                    else:
                        # 购物车结算
                        cur.execute("""
                            SELECT c.product_id,
                                c.sku_id,
                                c.quantity,
                                s.price,
                                p.is_member_product AS is_vip,
                                p.user_id as merchant_id,
                                c.specifications
                            FROM cart c
                            JOIN product_skus s ON s.id = c.sku_id
                            JOIN products p ON p.id = c.product_id
                            WHERE c.user_id = %s AND c.selected = 1
                        """, (user_id,))
                        items = cur.fetchall()
                        if not items:
                            return None

                        # 检查购物车中的商品是否属于同一商家
                        merchant_ids = set(item.get("merchant_id") or 0 for item in items)
                        if len(merchant_ids) > 1:
                            raise HTTPException(
                                status_code=400,
                                detail="购物车中包含不同商家的商品，请分开结算"
                            )
                        
                        # 如果未提供 merchant_id，从商品中推断
                        if merchant_id is None:
                            merchant_id = merchant_ids.pop() if merchant_ids else 0

                    # 确保 merchant_id 是整数
                    merchant_id = int(merchant_id or 0)

                    # ---------- 2. 优惠券商品类型验证（新增） ----------
                    has_vip = any(i["is_vip"] for i in items)

                    if coupon_id:
                        # 查询优惠券详情并锁定
                        cur.execute("""
                            SELECT id, amount, applicable_product_type, status, valid_from, valid_to, user_id
                            FROM coupons 
                            WHERE id = %s AND user_id = %s AND status = 'unused'
                            FOR UPDATE
                        """, (coupon_id, user_id))
                        coupon = cur.fetchone()

                        if not coupon:
                            raise HTTPException(status_code=400, detail="优惠券不存在、已被使用或不属于当前用户")

                        # 检查有效期
                        today = datetime.now().date()
                        if not (coupon['valid_from'] <= today <= coupon['valid_to']):
                            raise HTTPException(status_code=400, detail="优惠券不在有效期内")

                        # 检查商品类型匹配
                        applicable_type = coupon['applicable_product_type']
                        if applicable_type == 'member_only' and not has_vip:
                            raise HTTPException(status_code=400, detail="该优惠券仅限会员商品使用")
                        if applicable_type == 'normal_only' and has_vip:
                            raise HTTPException(status_code=400, detail="该优惠券仅限普通商品使用")

                        # 检查优惠券金额是否超过订单金额
                        coupon_amount = Decimal(str(coupon['amount']))
                        total_amount = sum(Decimal(str(i["quantity"])) * Decimal(str(i["price"])) for i in items)
                        if coupon_amount > total_amount:
                            raise HTTPException(status_code=400, detail="优惠券金额不能大于订单金额")

                    # ---------- 3. 地址信息 ----------
                    if delivery_way == "pickup":
                        consignee_name = consignee_phone = province = city = district = shipping_address = ""
                    elif custom_addr:
                        consignee_name = custom_addr.get("consignee_name")
                        consignee_phone = custom_addr.get("consignee_phone")
                        province = custom_addr.get("province", "")
                        city = custom_addr.get("city", "")
                        district = custom_addr.get("district", "")
                        shipping_address = custom_addr.get("detail", "")
                    else:
                        raise HTTPException(status_code=422, detail="必须上传收货地址或选择自提")

                    # ---------- 4. 订单主表 ----------
                    total = sum(Decimal(str(i["quantity"])) * Decimal(str(i["price"])) for i in items)

                    # ==================== 优化：更安全的订单号生成 ====================
                    # 使用 16 位十六进制 UUID 代替原来的 6 位数字，极大降低碰撞概率
                    order_number = (
                            datetime.now().strftime("%Y%m%d%H%M%S") +
                            str(user_id) +
                            uuid.uuid4().hex[:16]  # 16位十六进制，比原来6位数字更安全
                    )

                    init_status = "pending_pay"

                    # 修改后的 INSERT 语句，包含 merchant_id 字段
                    cur.execute("""
                        INSERT INTO orders(
                            user_id, merchant_id, order_number, total_amount, original_amount, status, is_vip_item,
                            consignee_name, consignee_phone,
                            province, city, district, shipping_address, delivery_way,
                            pay_way, auto_recv_time, refund_reason, expire_at,
                            pending_points, pending_coupon_id)
                        VALUES (%s, %s, %s, %s, %s, %s, %s,
                                %s, %s, %s, %s, %s, %s, %s,
                                'wechat', %s, %s, %s, %s, %s)
                    """, (
                        user_id, merchant_id, order_number, total, total, init_status, has_vip,
                        consignee_name, consignee_phone,
                        province, city, district, shipping_address, delivery_way,
                        datetime.now() + timedelta(days=7),
                        specifications,
                        datetime.now() + timedelta(hours=12) if init_status == "pending_pay" else None,
                        points_to_use or Decimal('0'),
                        coupon_id
                    ))
                    oid = cur.lastrowid

                    # ---------- 5. 库存校验 & 扣减 ----------
                    structure = get_table_structure(cur, "product_skus")
                    has_stock_field = 'stock' in structure['fields']
                    stock_select = (
                        f"COALESCE({_quote_identifier('stock')}, 0) AS {_quote_identifier('stock')}"
                        if has_stock_field and 'stock' in structure['asset_fields']
                        else _quote_identifier('stock')
                    ) if has_stock_field else "0 AS stock"

                    for i in items:
                        cur.execute(f"SELECT {stock_select} FROM {_quote_identifier('product_skus')} WHERE id=%s",
                                    (i['sku_id'],))
                        result = cur.fetchone()
                        current_stock = result.get('stock', 0) if result else 0
                        if current_stock < i["quantity"]:
                            raise HTTPException(status_code=400, detail=f"SKU {i['sku_id']} 库存不足")

                    # ---------- 6. 写订单明细 ----------
                    for i in items:
                        cur.execute("""
                            INSERT INTO order_items(order_id, product_id, sku_id, quantity, unit_price, total_price)
                            VALUES (%s, %s, %s, %s, %s, %s)
                        """, (
                            oid, i["product_id"], i["sku_id"], i["quantity"],
                            i["price"], Decimal(str(i["quantity"])) * Decimal(str(i["price"]))
                        ))

                    # ---------- 7. 扣库存 ----------
                    if has_stock_field:
                        for i in items:
                            cur.execute("UPDATE product_skus SET stock = stock - %s WHERE id = %s",
                                        (i["quantity"], i['sku_id']))

                    # ---------- 8. 清空购物车（仅购物车结算场景） ----------
                    if not buy_now:
                        cur.execute("DELETE FROM cart WHERE user_id = %s AND selected = 1", (user_id,))

                    # ==================== 新增：记录幂等 Key 使用（如果提供了的话） ====================
                    if idempotency_key and redis_client:
                        used_key = f"order:idempotency:{idempotency_key}"
                        redis_client.setex(used_key, 86400, order_number)  # 24小时过期

                    conn.commit()
                    logger.info(f"订单创建成功: {order_number}, 用户: {user_id}, 商家: {merchant_id}")
                    return order_number

        finally:
            # ==================== 新增：无论成功与否都释放 Redis 锁 ====================
            if lock_acquired and redis_client:
                try:
                    redis_client.delete(lock_key)
                except Exception as e:
                    logger.error(f"释放 Redis 锁失败: {e}")

    @staticmethod
    def list_by_user(user_id: int, status: Optional[str] = None):
        """按用户查询订单列表，附带首件商品和规格字段。"""
        with get_conn() as conn:
            with conn.cursor() as cur:
                select_fields = OrderManager._build_orders_select(cur)
                sql = f"SELECT {select_fields} FROM orders WHERE user_id = %s"
                params = [user_id]
                if status:
                    sql += " AND status = %s"
                    params.append(status)
                sql += " ORDER BY created_at DESC"
                cur.execute(sql, tuple(params))
                orders = cur.fetchall()

                for o in orders:
                    cur.execute(
                        """
                        SELECT oi.*, p.name
                        FROM order_items oi
                        JOIN products p ON oi.product_id = p.id
                        WHERE oi.order_id = %s
                        LIMIT 1
                        """,
                        (o["id"],)
                    )
                    first_item = cur.fetchone()
                    o["first_product"] = first_item
                    o["specifications"] = o.get("refund_reason")

                return orders

    @staticmethod
    def list_by_merchant(
        merchant_id: int,
        status: Optional[str] = None,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        page: int = 1,
        page_size: int = 20
    ) -> Dict[str, Any]:
        """
        按商家查询订单列表（支持分页、状态筛选、时间范围筛选）
        
        Args:
            merchant_id: 商家ID（对应 users.id）
            status: 订单状态筛选
            start_date: 开始日期（格式：YYYY-MM-DD）
            end_date: 结束日期（格式：YYYY-MM-DD）
            page: 页码，从1开始
            page_size: 每页数量
            
        Returns:
            包含订单列表、分页信息、统计信息的字典
        """
        with get_conn() as conn:
            with conn.cursor() as cur:
                # 构建查询条件
                where_conditions = ["merchant_id = %s"]
                params = [merchant_id]
                
                if status:
                    where_conditions.append("status = %s")
                    params.append(status)
                
                if start_date:
                    where_conditions.append("DATE(created_at) >= %s")
                    params.append(start_date)
                
                if end_date:
                    where_conditions.append("DATE(created_at) <= %s")
                    params.append(end_date)
                
                where_clause = " AND ".join(where_conditions)
                
                # 查询总数量
                count_sql = f"SELECT COUNT(*) as total FROM orders WHERE {where_clause}"
                cur.execute(count_sql, tuple(params))
                total = cur.fetchone()["total"]
                
                # 查询总金额
                amount_sql = f"""
                    SELECT 
                        COALESCE(SUM(total_amount), 0) as total_amount,
                        COALESCE(SUM(CASE WHEN status = 'completed' THEN total_amount ELSE 0 END), 0) as completed_amount
                    FROM orders 
                    WHERE {where_clause}
                """
                cur.execute(amount_sql, tuple(params))
                amount_stats = cur.fetchone()
                
                # 查询订单列表（分页）
                select_fields = OrderManager._build_orders_select(cur)
                offset = (page - 1) * page_size
                sql = f"""
                    SELECT {select_fields} 
                    FROM orders 
                    WHERE {where_clause}
                    ORDER BY created_at DESC
                    LIMIT %s OFFSET %s
                """
                query_params = params + [page_size, offset]
                cur.execute(sql, tuple(query_params))
                orders = cur.fetchall()
                
                # 为每个订单查询商品明细和用户信息
                for o in orders:
                    # 查询商品明细
                    cur.execute(
                        """
                        SELECT oi.*, p.name as product_name, p.cover as product_cover
                        FROM order_items oi
                        JOIN products p ON oi.product_id = p.id
                        WHERE oi.order_id = %s
                        """,
                        (o["id"],)
                    )
                    o["items"] = cur.fetchall()
                    
                    # 查询用户信息
                    cur.execute(
                        "SELECT id, name, mobile, avatar FROM users WHERE id = %s",
                        (o["user_id"],)
                    )
                    user_info = cur.fetchone()
                    o["user_info"] = user_info
                
                return {
                    "list": orders,
                    "pagination": {
                        "page": page,
                        "page_size": page_size,
                        "total": total,
                        "total_pages": (total + page_size - 1) // page_size
                    },
                    "statistics": {
                        "total_amount": float(amount_stats["total_amount"]),
                        "completed_amount": float(amount_stats["completed_amount"]),
                        "order_count": total
                    }
                }

    @staticmethod
    def detail(order_number: str) -> Optional[dict]:
        """查询单个订单详情（含用户、地址、商品明细、商家信息）。"""
        with get_conn() as conn:
            with conn.cursor() as cur:
                select_fields = OrderManager._build_orders_select(cur)
                cur.execute(
                    f"SELECT {select_fields} FROM orders WHERE order_number=%s LIMIT 1",
                    (order_number,)
                )
                order = cur.fetchone()
                if not order:
                    return None

                order_id = order.get("id")
                user_id = order.get("user_id")
                merchant_id = order.get("merchant_id") or 0

                # 商品明细
                cur.execute(
                    """
                    SELECT oi.*, p.name AS product_name, p.is_member_product, p.cover AS product_cover
                    FROM order_items oi
                    LEFT JOIN products p ON oi.product_id = p.id
                    WHERE oi.order_id = %s
                    """,
                    (order_id,)
                )
                items = cur.fetchall()

                # 用户信息
                user_info = None
                if user_id:
                    cur.execute(
                        "SELECT id, name, mobile, avatar, member_level, member_points FROM users WHERE id=%s",
                        (user_id,)
                    )
                    user_info = cur.fetchone()

                # 商家信息
                merchant_info = None
                if merchant_id and merchant_id > 0:
                    cur.execute(
                        """
                        SELECT u.id, u.name, u.mobile, u.avatar, u.wechat_sub_mchid,
                               ms.store_name, ms.store_logo_image_id, ms.store_address
                        FROM users u
                        LEFT JOIN merchant_stores ms ON ms.user_id = u.id
                        WHERE u.id = %s AND u.is_merchant = 1
                        """,
                        (merchant_id,)
                    )
                    merchant_info = cur.fetchone()

                # 地址信息直接取订单中的收货字段
                address = {
                    "consignee_name": order.get("consignee_name"),
                    "consignee_phone": order.get("consignee_phone"),
                    "province": order.get("province"),
                    "city": order.get("city"),
                    "district": order.get("district"),
                    "detail": order.get("shipping_address"),
                }

                return {
                    "order_info": order,
                    "user": user_info,
                    "merchant": merchant_info,
                    "address": address,
                    "items": items,
                    "specifications": order.get("refund_reason"),
                }

    @staticmethod
    def update_status(order_number: str, new_status: str, reason: Optional[str] = None,
                      external_conn=None) -> bool:
        """统一的订单状态更新，支持外部连接复用。"""

        def _apply_update(cur) -> bool:
            cur.execute("SHOW COLUMNS FROM orders")
            cols = {row.get("Field") for row in cur.fetchall()}

            updates = ["status=%s", "updated_at=NOW()"]
            params: List[Any] = [new_status]

            # 需要记录原因时，优先写入 status_reason / remark，避免覆盖 refund_reason
            if reason:
                for col in ("status_reason", "remark"):
                    if col in cols:
                        updates.append(f"{col}=%s")
                        params.append(reason)
                        break

            if new_status in ("pending_ship", "pending_recv") and "paid_at" in cols:
                updates.append("paid_at=COALESCE(paid_at, NOW())")
            if new_status == "pending_recv" and "shipped_at" in cols:
                updates.append("shipped_at=COALESCE(shipped_at, NOW())")
            if new_status == "completed" and "completed_at" in cols:
                updates.append("completed_at=COALESCE(completed_at, NOW())")

            params.append(order_number)
            cur.execute(f"UPDATE orders SET {', '.join(updates)} WHERE order_number=%s", tuple(params))
            return cur.rowcount > 0

        if external_conn:
            cur = external_conn.cursor()
            try:
                return _apply_update(cur)
            finally:
                cur.close()
        else:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    updated = _apply_update(cur)
                    conn.commit()
                    return updated

    # ==================== 新增：确认收货处理（前端调用微信组件后回调） ====================
    @staticmethod
    def confirm_receive(order_number: str, user_id: Optional[int] = None,
                        wx_confirm_result: Optional[Dict] = None) -> Dict[str, Any]:
        """
        用户确认收货（前端调用微信确认收货组件成功后回调）

        修复：增加重试机制处理微信状态同步延迟（最多3次，间隔1.5秒）
        """
        result = {
            "ok": False,
            "message": "",
            "wx_verified": False
        }

        with get_conn() as conn:
            with conn.cursor() as cur:
                # 1. 查询订单信息
                cur.execute(
                    """SELECT id, user_id, status, transaction_id, order_number, total_amount
                       FROM orders WHERE order_number=%s""",
                    (order_number,)
                )
                order = cur.fetchone()

                if not order:
                    result["message"] = "订单不存在"
                    return result

                # 验证用户权限（如果提供了user_id）
                if user_id and order['user_id'] != user_id:
                    result["message"] = "无权操作该订单"
                    return result

                if order['status'] != 'pending_recv':
                    result["message"] = f"订单状态不正确，当前状态：{order['status']}"
                    return result

                transaction_id = order.get('transaction_id')
                if not transaction_id:
                    result["message"] = "缺少微信支付单号，无法确认收货"
                    return result

                # 可选：验证前端传入的微信组件结果
                if wx_confirm_result and wx_confirm_result.get('order_id') != transaction_id:
                    logger.warning(f"[confirm_receive] 前端传入的transaction_id与订单不符")
                    result["message"] = "验证失败：订单信息不匹配"
                    return result

                # ===== 核心修复：带重试的微信状态查询 =====
                max_retries = 3
                retry_delay = 1.5  # 秒
                wx_result = None
                order_state = None

                for attempt in range(max_retries):
                    try:
                        wx_result = WechatShippingManager.get_order(transaction_id)

                        # 修复：errcode 转为字符串比较
                        if str(wx_result.get('errcode', '')) != '0':
                            error_msg = wx_result.get('errmsg', '未知错误')
                            result["message"] = f"查询微信订单状态失败：{error_msg}"
                            return result

                        # 修复：兼容 order_state 在顶层或嵌套在 order 对象中的情况，并转为整数
                        raw_state = wx_result.get('order_state') or (wx_result.get('order') or {}).get('order_state')
                        try:
                            order_state = int(raw_state) if raw_state is not None else None
                        except (ValueError, TypeError):
                            order_state = None
                            logger.warning(f"[confirm_receive] order_state 转换失败，原始值: {raw_state}")

                        # 已确认收货(3)或交易完成(4)，立即通过
                        if order_state in (3, 4):
                            break

                        # 【关键修复】只要不是3或4（包括None、1、2、5、6），且还有重试次数，就等待后重试
                        if attempt < max_retries - 1:
                            logger.info(
                                f"[confirm_receive] 订单 {order_number} 微信状态 {order_state}，"
                                f"第 {attempt + 1}/{max_retries} 次重试，等待 {retry_delay}s..."
                            )
                            time.sleep(retry_delay)
                        else:
                            # 最后一次了，必须break，否则死循环
                            break

                    except Exception as e:
                        logger.error(f"[confirm_receive] 查询微信状态异常: {e}")
                        if attempt == max_retries - 1:
                            result["message"] = "校验微信收货状态失败，请稍后重试"
                            return result
                        time.sleep(retry_delay)

                # 状态校验
                if order_state == 3:
                    result["wx_verified"] = True
                    verify_msg = "微信已确认收货"
                elif order_state == 4:
                    result["wx_verified"] = True
                    verify_msg = "微信交易已完成"
                else:
                    # 状态映射表
                    state_map = {1: "待发货", 2: "已发货未收货", 5: "已退款", 6: "资金待结算"}

                    if order_state is None:
                        result["message"] = "微信状态尚未同步，请稍等2-3分钟后重试"
                        logger.warning(f"[confirm_receive] 订单 {order_number} 微信返回状态为None")
                    else:
                        current_state_name = state_map.get(order_state, f"异常状态({order_state})")
                        result["message"] = (
                            f"微信端未确认收货，当前状态：{current_state_name}。"
                            f"请确保在微信小程序内点击确认收货按钮。"
                        )
                    return result

                # 更新订单状态
                updated = OrderManager.update_status(
                    order_number,
                    "completed",
                    f"用户确认收货({verify_msg})",
                    external_conn=conn
                )

                if not updated:
                    result["message"] = "更新订单状态失败"
                    return result

                # 记录确认收货日志（这里是你缺失的代码）
                try:
                    cur.execute("""
                        INSERT INTO wechat_shipping_logs 
                        (order_id, order_number, transaction_id, action_type, is_success, remark, response_data, created_at)
                        VALUES (%s, %s, %s, 'sync', 1, %s, %s, NOW())
                    """, (
                        order['id'],
                        order_number,
                        transaction_id,
                        f"用户确认收货, 微信验证: {result['wx_verified']}",
                        json.dumps(wx_result)
                    ))
                    conn.commit()
                except Exception as e:
                    logger.error(f"[confirm_receive] 记录日志失败: {e}")
                    # 日志记录失败不影响主流程，但最好回滚或记录
                    conn.rollback()

                result["ok"] = True
                result["message"] = f"确认收货成功（{verify_msg}），资金将在微信侧结算"

                # 可选：触发后续业务（如积分发放）
                try:
                    # 如果需要在确认收货时触发某些业务逻辑，可以在这里调用
                    pass
                except Exception as e:
                    logger.error(f"[confirm_receive] 订单 {order_number} 后续处理异常: {e}")

                return result

    @staticmethod
    def export_to_excel(order_numbers: List[str]) -> bytes:
        """
        导出订单详情（包含资金拆分明细）
        生成两个工作表：订单详情、资金拆分明细
        """
        # 账户类型中英文映射
        account_type_map = {
            "merchant_balance": "商家余额",
            "public_welfare": "公益基金",
            "maintain_pool": "平台维护",
            "subsidy_pool": "周补贴池",
            "director_pool": "联创奖励",
            "shop_pool": "社区店",
            "city_pool": "城市运营中心",
            "branch_pool": "大区分公司",
            "fund_pool": "事业发展基金",
            "company_points": "公司积分账户",
            "company_balance": "公司余额账户",
            "platform_revenue_pool": "平台收入池（会员商品）",
            "wx_applyment_fee": "微信进件手续费",
            "income": "收入",
            "expense": "支出"
        }

        wb = Workbook()

        # ========== 第一个工作表：订单详情 ==========
        ws1 = wb.active
        ws1.title = "订单详情"

        headers1 = [
            "订单号", "商家ID", "商家名称", "订单状态", "总金额", "原始金额", "积分抵扣", "实付金额",
            "支付方式", "配送方式", "是否会员订单",
            "用户ID", "用户姓名", "用户手机号",
            "收货人", "收货电话", "省份", "城市", "区县", "详细地址",
            "商品信息", "商品规格", "下单时间", "支付时间", "发货时间"
        ]

        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col_idx, header in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # ========== 第二个工作表：资金拆分明细 ==========
        ws2 = wb.create_sheet(title="资金拆分")
        headers2 = [
            "订单号", "商家ID", "账户类型", "变动金额", "变动后余额",
            "流水类型", "备注", "创建时间"
        ]

        for col_idx, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 查询数据并填充
        row_idx1 = 2
        row_idx2 = 2

        with get_conn() as conn:
            with conn.cursor() as cur:
                for order_number in order_numbers:
                    # 1. 查询订单详情
                    order_data = OrderManager.detail(order_number)
                    if not order_data:
                        continue

                    order_info = order_data["order_info"]
                    user_info = order_data["user"]
                    merchant_info = order_data.get("merchant")
                    address = order_data["address"] or {}
                    items = order_data["items"]
                    specifications = order_data.get("specifications") or {}

                    # 计算实付金额（用于计算20%平台抽成）
                    total = Decimal(str(order_info.get("total_amount", 0)))
                    points_discount = Decimal(str(order_info.get("points_discount", 0)))
                    actual_pay = total - points_discount

                    # 商品信息拼接
                    product_info = "\n".join([
                        f"{item.get('product_name', '')} x{item.get('quantity', 0)} @¥{item.get('unit_price', 0)}"
                        for item in items
                    ])

                    # 规格信息
                    spec_str = ""
                    if isinstance(specifications, dict):
                        spec_str = "\n".join([f"{k}: {v}" for k, v in specifications.items()])

                    # 判断发货时间（自提订单=支付时间）
                    shipped_at = order_info.get("shipped_at", "")
                    if order_info.get("delivery_way") == "pickup":
                        shipped_at = order_info.get("paid_at", "")

                    # 填充订单详情行
                    row_data1 = [
                        order_info.get("order_number", ""),
                        order_info.get("merchant_id", 0),
                        merchant_info.get("store_name") or merchant_info.get("name", "平台自营") if merchant_info else "平台自营",
                        order_info.get("status", ""),
                        float(total),
                        float(order_info.get("original_amount", 0)),
                        float(points_discount),
                        float(actual_pay),
                        order_info.get("pay_way", "wechat"),
                        order_info.get("delivery_way", "platform"),
                        "是" if order_info.get("is_member_order") else "否",
                        user_info.get("id", ""),
                        user_info.get("name", ""),
                        user_info.get("mobile", ""),
                        address.get("consignee_name", ""),
                        address.get("consignee_phone", ""),
                        address.get("province", ""),
                        address.get("city", ""),
                        address.get("district", ""),
                        address.get("detail", ""),
                        product_info,
                        spec_str,
                        order_info.get("created_at", ""),
                        order_info.get("paid_at", ""),
                        shipped_at  # ← 使用修改后的发货时间
                    ]

                    for col_idx, value in enumerate(row_data1, 1):
                        cell = ws1.cell(row=row_idx1, column=col_idx, value=value)
                        cell.alignment = Alignment(vertical="center", wrap_text=True)
                        cell.border = thin_border
                        if col_idx in [5, 6, 7, 8]:  # 金额列
                            cell.number_format = '¥#,##0.00'

                    row_idx1 += 1

                    # 2. 查询资金拆分明细（account_flow表）
                    cur.execute("""
                        SELECT account_type, change_amount, balance_after, 
                               flow_type, remark, created_at
                        FROM account_flow 
                        WHERE remark LIKE %s
                        ORDER BY created_at ASC
                    """, (f"%{order_number}%",))

                    flows = cur.fetchall()

                    # 计算平台抽成总额（20%）
                    platform_fee = float(actual_pay) * 0.2

                    for flow in flows:
                        account_type_en = flow.get("account_type", "")

                        # 特殊处理：商家余额显示为X雨点（20%）
                        if account_type_en == "merchant_balance":
                            display_amount = f"{int(platform_fee)}雨点"
                            account_type_cn = "商家余额"  # ← 改回"商家余额"
                            balance_after_display = "-"
                            is_platform_fee_row = True
                        else:
                            # 其他账户类型正常显示
                            account_type_cn = account_type_map.get(account_type_en, account_type_en)
                            display_amount = float(flow.get("change_amount", 0))
                            balance_after_display = float(flow.get("balance_after", 0))
                            is_platform_fee_row = False

                        row_data2 = [
                            order_number,
                            order_info.get("merchant_id", 0),
                            account_type_cn,
                            display_amount,
                            balance_after_display,
                            flow.get("flow_type", ""),
                            flow.get("remark", ""),
                            flow.get("created_at", "")
                        ]

                        for col_idx, value in enumerate(row_data2, 1):
                            cell = ws2.cell(row=row_idx2, column=col_idx, value=value)
                            cell.alignment = Alignment(vertical="center")
                            cell.border = thin_border

                            # 只有非商家余额行才设置货币格式
                            if col_idx in [4, 5] and not is_platform_fee_row:
                                cell.number_format = '¥#,##0.00'

                        row_idx2 += 1

        # 调整列宽（订单详情表）
        for column in ws1.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws1.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # 调整列宽（资金拆分表）
        for column in ws2.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws2.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # 保存到内存
        excel_data = BytesIO()
        wb.save(excel_data)
        excel_data.seek(0)
        return excel_data.getvalue()


# ---------------- 请求模型 ----------------
class DeliveryWay(str, Enum):
    platform = "platform"  # 平台配送
    pickup = "pickup"


class OrderCreate(BaseModel):
    user_id: int
    delivery_way: DeliveryWay = DeliveryWay.platform  # 新增
    address_id: Optional[int] = None
    custom_address: Optional[dict] = None
    specifications: Optional[str] = None
    buy_now: bool = False
    buy_now_items: Optional[List[Dict[str, Any]]] = None
    points_to_use: Optional[Decimal] = Decimal('0')  # 新增：积分抵扣数量
    coupon_id: Optional[int] = None  # 新增：优惠券ID
    # ==================== 新增：幂等性 Key（前端生成 UUID） ====================
    idempotency_key: Optional[str] = None  # 用于防止重复提交
    # ==================== 新增：商家ID ====================
    merchant_id: Optional[int] = None  # 商家ID，不传则自动从商品推断


class OrderPay(BaseModel):
    order_number: str
    pay_way: str
    coupon_id: Optional[int] = None
    points_to_use: Optional[Decimal] = Decimal('0')


class StatusUpdate(BaseModel):
    order_number: str
    new_status: str
    reason: Optional[str] = None


class WechatPayParams(BaseModel):
    appId: str
    timeStamp: str
    nonceStr: str
    package: str
    signType: str
    paySign: str


# ==================== 新增：确认收货请求模型（修改后） ====================
class ConfirmReceiveRequest(BaseModel):
    order_number: str
    # 可选：微信组件返回的确认结果（用于增强验证）
    wx_confirm_result: Optional[Dict[str, Any]] = None


# ==================== 新增：商家订单查询请求模型 ====================
class MerchantOrdersQuery(BaseModel):
    status: Optional[str] = None
    start_date: Optional[str] = None  # 格式：YYYY-MM-DD
    end_date: Optional[str] = None    # 格式：YYYY-MM-DD
    page: int = 1
    page_size: int = 20


# ---------------- 路由 ----------------
@router.post("/create", summary="创建订单")
def create_order(body: OrderCreate):
    """
    创建订单接口（已增加幂等性校验，支持多商家订单）
    
    - 请前端在提交时生成 idempotency_key（UUID），用于防止重复创建
    - 如果同一用户 1 分钟内已有未取消订单，会返回错误
    - 如果 Redis 锁未释放（5 秒内），会返回 429 错误
    - 如果 buy_now_items 中的商品属于不同商家，会返回错误（一笔订单只能属于一个商家）
    - 如果不传 merchant_id，系统会自动从商品中推断商家
    """
    no = OrderManager.create(
        body.user_id,
        body.address_id,
        body.custom_address,
        specifications=body.specifications,
        buy_now=body.buy_now,
        buy_now_items=body.buy_now_items,
        delivery_way=body.delivery_way,
        points_to_use=body.points_to_use,  # 传递积分参数
        coupon_id=body.coupon_id,  # 传递优惠券参数
        idempotency_key=body.idempotency_key,  # 传递幂等 Key
        merchant_id=body.merchant_id  # 传递商家ID
    )
    if not no:
        raise HTTPException(status_code=422, detail="购物车为空或地址缺失")
    return {"order_number": no}


@router.get("/{user_id}", summary="查询用户订单列表")
def list_orders(user_id: int, status: Optional[str] = None):
    return OrderManager.list_by_user(user_id, status)


@router.get("/detail/{order_number}", summary="查询订单详情")
def order_detail(order_number: str):
    d = OrderManager.detail(order_number)
    if not d:
        raise HTTPException(status_code=404, detail="订单不存在")
    return d


@router.post("/status", summary="更新订单状态")
def update_status(body: StatusUpdate):
    return {"ok": OrderManager.update_status(body.order_number, body.new_status, body.reason)}


# ==================== 新增：按商家查询订单接口 ====================
@router.get("/merchant/{merchant_id}", summary="查询商家订单列表")
def list_merchant_orders(
    merchant_id: int,
    status: Optional[str] = Query(None, description="订单状态筛选"),
    start_date: Optional[str] = Query(None, description="开始日期(YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="结束日期(YYYY-MM-DD)"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页数量")
):
    """
    根据商家ID查询订单列表（支持分页、状态筛选、时间范围筛选）
    
    返回内容包括：
    - list: 订单列表（包含商品明细和用户信息）
    - pagination: 分页信息
    - statistics: 统计信息（订单总金额、已完成金额、订单数量）
    
    权限说明：此接口仅返回该商家的订单，需要配合权限校验使用
    """
    result = OrderManager.list_by_merchant(
        merchant_id=merchant_id,
        status=status,
        start_date=start_date,
        end_date=end_date,
        page=page,
        page_size=page_size
    )
    return result


# ==================== 新增：确认收货接口（前端调用微信组件后回调） ====================
@router.post("/confirm-receive", summary="用户确认收货（微信组件回调）")
def confirm_receive(body: ConfirmReceiveRequest):
    """
    用户确认收货接口

    调用时机：前端调用微信确认收货组件(wx.confirmOrderReceiption)成功后回调

    **重要**：必须通过微信组件完成确认，否则资金无法结算！

    前置条件：
    1. 订单状态必须为 pending_recv（已发货待收货）
    2. 【强制】微信侧订单状态必须为已确认收货或交易完成

    业务流程：
    1. 验证订单状态和权限
    2. 强制查询微信侧状态（必须为3或4）
    3. 更新订单状态为 completed（已完成）
    4. 触发资金结算（如果在支付时未结算）

    失败情况：
    - 如果微信侧未确认收货，会返回错误，前端应引导用户去小程序订单列表确认收货
    """
    result = OrderManager.confirm_receive(
        body.order_number,
        wx_confirm_result=body.wx_confirm_result
    )

    if not result["ok"]:
        raise HTTPException(status_code=400, detail=result["message"])

    return result


def auto_receive_task(db_cfg: dict = None):
    """自动收货守护进程（不再发放积分）"""
    import threading
    import time
    from datetime import datetime

    def run():
        while True:
            try:
                from core.database import get_conn
                with get_conn() as conn:
                    with conn.cursor() as cur:
                        now = datetime.now()
                        cur.execute(
                            "SELECT id, order_number, total_amount FROM orders "
                            "WHERE status='pending_recv' AND auto_recv_time<=%s",
                            (now,)
                        )
                        for row in cur.fetchall():
                            order_id = row["id"]
                            order_number = row["order_number"]

                            # 更新订单状态为已完成
                            cur.execute(
                                "UPDATE orders SET status='completed' WHERE id=%s",
                                (order_id,)
                            )

                            # ✅ **移除**：积分已在支付时发放，自动收货不再发放
                            # try:
                            #     fs = FinanceService()
                            #     fs.grant_points_on_receive(order_number, external_conn=conn)
                            # except Exception as e:
                            #     logger.error(f"[auto_receive] 订单{order_number}积分发放失败: {e}", exc_info=True)

                            conn.commit()
                            logger.debug(f"[auto_receive] 订单 {order_number} 已自动完成。")
            except Exception as e:
                logger.error(f"[auto_receive] 异常: {e}")
            time.sleep(3600)  # 每小时检查一次

    t = threading.Thread(target=run, daemon=True)
    t.start()
    logger.info("自动收货守护进程已启动（不再发放积分）")


class OrderExportRequest(BaseModel):
    order_numbers: List[str]


class OrderExportByTimeRequest(BaseModel):
    start_time: str  # 格式：2025-01-01 00:00:00
    end_time: str  # 格式：2025-01-31 23:59:59
    status: Optional[str] = None  # 可选：按订单状态筛选（如 pending_ship, completed 等）


@router.post("/export", summary="导出订单详情到Excel")
def export_orders(body: OrderExportRequest):
    """
    批量导出订单详情为Excel文件
    请求示例: {"order_numbers": ["20250101120000", "20250101120001"]}
    """
    if not body.order_numbers:
        raise HTTPException(status_code=422, detail="订单号列表不能为空")

    # 限制一次最多导出1000个订单
    if len(body.order_numbers) > 1000:
        raise HTTPException(status_code=422, detail="单次导出订单数不能超过1000个")

    try:
        excel_data = OrderManager.export_to_excel(body.order_numbers)
        return StreamingResponse(
            BytesIO(excel_data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=orders_export.xlsx"}
        )
    except Exception as e:
        logger.error(f"导出订单失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@router.post("/export/by-time", summary="按时间范围导出订单")
def export_orders_by_time(body: OrderExportByTimeRequest):
    """
    按时间范围批量导出订单详情
    请求示例: {
        "start_time": "2025-01-01 00:00:00",
        "end_time": "2025-01-31 23:59:59",
        "status": "completed"  // 可选，不填则导出所有状态
    }
    """
    # 1. 时间格式校验
    try:
        start = datetime.strptime(body.start_time, "%Y-%m-%d %H:%M:%S")
        end = datetime.strptime(body.end_time, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        raise HTTPException(
            status_code=422,
            detail="时间格式错误，请使用：YYYY-MM-DD HH:MM:SS"
        )

    if end < start:
        raise HTTPException(
            status_code=422,
            detail="结束时间不能早于开始时间"
        )

    # 2. 限制时间范围（防止导出数据量过大，最多31天）
    if (end - start).days > 31:
        raise HTTPException(
            status_code=422,
            detail="时间范围不能超过31天"
        )

    # 3. 查询时间范围内的订单号
    with get_conn() as conn:
        with conn.cursor() as cur:
            if body.status:
                # 按状态筛选
                sql = """
                    SELECT order_number 
                    FROM orders 
                    WHERE created_at >= %s 
                      AND created_at <= %s 
                      AND status = %s 
                    ORDER BY created_at DESC 
                    LIMIT 500
                """
                cur.execute(sql, (body.start_time, body.end_time, body.status))
            else:
                # 不筛选状态，导出全部
                sql = """
                    SELECT order_number 
                    FROM orders 
                    WHERE created_at >= %s 
                      AND created_at <= %s 
                    ORDER BY created_at DESC 
                    LIMIT 500
                """
                cur.execute(sql, (body.start_time, body.end_time))

            rows = cur.fetchall()
            order_numbers = [row["order_number"] for row in rows]

    # 4. 检查是否有数据
    if not order_numbers:
        raise HTTPException(
            status_code=404,
            detail="该时间段内没有符合条件的订单"
        )

    # 5. 调用现有的导出方法生成Excel
    try:
        excel_data = OrderManager.export_to_excel(order_numbers)

        # 生成带时间范围的文件名
        start_str = body.start_time[:10].replace("-", "")
        end_str = body.end_time[:10].replace("-", "")
        filename = f"orders_{start_str}_to_{end_str}.xlsx"
        if body.status:
            filename = f"orders_{body.status}_{start_str}_to_{end_str}.xlsx"

        return StreamingResponse(
            BytesIO(excel_data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error(f"按时间导出订单失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


# ==================== 修改：模块导入时启动新增的后台任务 ====================
start_order_expire_task()
# 新增：启动微信状态同步任务
start_wechat_status_sync_task()